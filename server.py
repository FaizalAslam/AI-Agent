from flask import Flask, render_template, request, jsonify
import threading
import json
import time
import logging
import webbrowser
import traceback
import os
import re

# ---- Core modules ---------------------------------------------------------
from modules import system_core, ui, config

# ---- Office Agent (Project 2) --------------------------------------------
from utils.command_buffer import CommandBuffer
from utils import command_map
from executor.excel_executor import ExcelExecutor
from executor.word_executor import WordExecutor
from executor.ppt_executor import PowerPointExecutor
from parser.command_parser import parse_command
from ai.openai_handler import OpenAIHandler
from listener.keyboard_listener import KeyboardListener
from listener.clipboard_listener import ClipboardListener
try:
    from listener.voice_listener import VoiceListener
    VOICE_MODULE_AVAILABLE = True
except Exception:
    VoiceListener = None
    VOICE_MODULE_AVAILABLE = False

# ---- Optional modules (graceful fallback) --------------------------------
try:
    from modules import ocr_utils
    OCR_AVAILABLE = True
except Exception as e:
    print(f"OCR unavailable: {e}")
    OCR_AVAILABLE = False

try:
    from modules import pdf_utils
    PDF_AVAILABLE = True
except Exception as e:
    print(f"PDF unavailable: {e}")
    PDF_AVAILABLE = False

try:
    from modules import pdf_reader
    READER_AVAILABLE = True
except Exception as e:
    print(f"PDF Reader unavailable: {e}")
    READER_AVAILABLE = False

try:
    from modules import gui_automation
    GUI_AVAILABLE = True
except Exception as e:
    print(f"GUI unavailable: {e}")
    GUI_AVAILABLE = False

try:
    from modules import pdf_editor
    PDF_EDITOR_AVAILABLE = True
except Exception as e:
    print(f"PDF Editor unavailable: {e}")
    PDF_EDITOR_AVAILABLE = False

try:
    import keyboard
    KEYBOARD_AVAILABLE = True
except ImportError:
    print("keyboard not found — pip install keyboard")
    KEYBOARD_AVAILABLE = False

# ---- Logging --------------------------------------------------------------
logging.basicConfig(
    filename="agent.log",
    level=logging.INFO,
    format="%(asctime)s - %(message)s",
    datefmt="%H:%M:%S",
    filemode="w"
)
logging.getLogger("werkzeug").setLevel(logging.ERROR)

# ---- Flask app ------------------------------------------------------------
app = Flask(__name__)

# ---- Shared state ---------------------------------------------------------
last_ocr = {"text": "", "pending": False}

# ---- Office Agent setup ---------------------------------------------------
OFFICE_APPS = {"excel", "word", "powerpoint", "ppt"}
OFFICE_OUTPUTS = {
    "excel": "output.xlsx",
    "word": "output.docx",
    "powerpoint": "output.pptx",
    "ppt": "output.pptx",
}
OFFICE_DEPENDENCIES = {
    "excel": ("openpyxl", "openpyxl"),
    "word": ("docx", "python-docx"),
    "powerpoint": ("pptx", "python-pptx"),
    "ppt": ("pptx", "python-pptx"),
}
_cmd_buf = CommandBuffer()
_clipboard_listener = ClipboardListener(_cmd_buf)
_keyboard_listener = KeyboardListener(_handle_global_command := None, _cmd_buf)
_voice_listener = VoiceListener(_handle_global_command) if VOICE_MODULE_AVAILABLE else None
voice_state = {"enabled": False}
_openai_handler = OpenAIHandler()


# ---- Office Agent helpers -------------------------------------------------
def _safe_speak(text):
    try:
        ui.speak(text)
    except Exception:
        pass


def _extract_office_agent_command(raw_text):
    text = (raw_text or "").strip()
    match = re.match(r"^agent\s*:\s*(excel|word|powerpoint|ppt)\s*:\s*(.+)$", text, re.IGNORECASE)
    if not match:
        return None, None
    app_name = match.group(1).lower().strip()
    command_text = match.group(2).strip()
    return app_name, command_text


def _resolve_actions(app_name, command_text):
    def _estimate_subcommands(text):
        protected = re.sub(
            r"(\d+\s*(?:col|cols|column|columns)\s+)and(\s*\d+\s*(?:row|rows))",
            r"\1__AND__\2",
            (text or "").lower().strip()
        )
        protected = re.sub(
            r"(\d+\s*(?:row|rows)\s+)and(\s*\d+\s*(?:col|cols|column|columns))",
            r"\1__AND__\2",
            protected
        )
        parts = re.split(r"\s+(?:and|then|also|after that|next)\s+", protected)
        return max(1, len([p for p in parts if p.strip()]))

    def _actions_cover_command_intents(app, text, actions):
        low = (text or "").lower()
        names = {
            str(a.get("action", "")).strip().lower()
            for a in (actions or [])
            if isinstance(a, dict)
        }
        if not names:
            return False

        checks = []
        if app == "excel":
            if "background color" in low:
                checks.append("set_bg_color" in names)
            if "font color" in low:
                checks.append("set_font_color" in names)
            if "font size" in low:
                checks.append("set_font_size" in names)
            if "number format" in low:
                checks.append("set_number_format" in names)
            if "formula" in low or "sum" in low:
                checks.append("write_formula" in names)
            if "rename" in low and "sheet" in low:
                checks.append("rename_sheet" in names)
            if "insert row" in low:
                checks.append("insert_row" in names)
            if "[[" in low and "write" in low and "values" in low:
                checks.append("write_range" in names)

            # If command explicitly targets multiple cells for background color,
            # ensure cached actions include multiple bg actions.
            if "background color" in low and "cells" in low and re.search(r"\b[A-Z]{1,3}\d{1,7}\b.*\band\b.*\b[A-Z]{1,3}\d{1,7}\b", text, re.IGNORECASE):
                bg_count = sum(1 for a in (actions or []) if isinstance(a, dict) and str(a.get("action", "")).lower() == "set_bg_color")
                checks.append(bg_count >= 2)

        return all(checks) if checks else True

    cache_key, cached_actions, cache_score = command_map.get_cached_actions(app_name, command_text)
    # Use cache only for exact matches; fuzzy cache reuse can apply stale actions
    # to similar-but-different commands.
    if cached_actions and cache_score == 100:
        cached_count = len([a for a in cached_actions if isinstance(a, dict) and a.get("action")])
        clause_count = _estimate_subcommands(command_text)
        if cached_count >= clause_count and _actions_cover_command_intents(app_name, command_text, cached_actions):
            logging.info(f"Office cache hit [{app_name}] score={cache_score}: {command_text}")
            return cache_key or command_text, cached_actions, "command-cache"
        logging.info(
            f"Ignoring stale cache for [{app_name}] command (cached={cached_count}, clauses={clause_count}): {command_text}"
        )

    actions = parse_command(app_name, command_text)
    if actions:
        # If parser returns fewer actions than apparent command clauses,
        # try API and prefer the richer valid result.
        clause_count = _estimate_subcommands(command_text)
        if clause_count > len(actions):
            ai_actions = _openai_handler.interpret(app_name, command_text)
            if isinstance(ai_actions, dict):
                ai_actions = [ai_actions]
            if isinstance(ai_actions, list) and ai_actions:
                normalized_ai = [a for a in ai_actions if isinstance(a, dict) and a.get("action")]
                if len(normalized_ai) >= len(actions):
                    command_map.save_actions(app_name, command_text, normalized_ai)
                    return command_text, normalized_ai, "openai-fallback"
        command_map.save_actions(app_name, command_text, actions)
        return command_text, actions, "json-parser"

    ai_actions = _openai_handler.interpret(app_name, command_text)
    if isinstance(ai_actions, dict):
        ai_actions = [ai_actions]
    if isinstance(ai_actions, list) and ai_actions:
        normalized = [a for a in ai_actions if isinstance(a, dict) and a.get("action")]
        if normalized:
            command_map.save_actions(app_name, command_text, normalized)
            return command_text, normalized, "openai-fallback"

    return command_text, [], "no-match"


def _extract_named_file_path(command_text, app_name):
    text = (command_text or "").strip()
    ext = {
        "excel": "xlsx",
        "word": "docx",
        "powerpoint": "pptx",
        "ppt": "pptx",
    }.get(app_name)
    if not text or not ext:
        return ""

    def _sanitize_base(name):
        cleaned = re.sub(r'[<>:"/\\|?*]+', "", (name or "").strip())
        cleaned = re.split(r"\s+(?:and|then|with|in|on)\b", cleaned, maxsplit=1, flags=re.IGNORECASE)[0]
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
        return cleaned

    quoted = re.search(r'["\']([^"\']+\.' + re.escape(ext) + r')["\']', text, re.IGNORECASE)
    if quoted:
        return os.path.abspath(quoted.group(1).strip())

    plain = re.search(r'\b([A-Za-z0-9_\- .]+\.' + re.escape(ext) + r')\b', text, re.IGNORECASE)
    if plain:
        return os.path.abspath(plain.group(1).strip())

    # Support "named demo" or "called demo" without extension.
    named = re.search(r'\b(?:named|called|name)\s*[:=]?\s*["\']?([A-Za-z0-9_\- ]{1,100})["\']?\b', text, re.IGNORECASE)
    if named:
        base = _sanitize_base(named.group(1))
        if base:
            return os.path.abspath(f"{base}.{ext}")

    return ""


def _next_available_path(path):
    base, ext = os.path.splitext(os.path.abspath(path))
    candidate = f"{base}{ext}"
    idx = 1
    while os.path.exists(candidate):
        candidate = f"{base}_{idx}{ext}"
        idx += 1
    return candidate


def _generate_new_output_path(app_name):
    ext = {
        "excel": "xlsx",
        "word": "docx",
        "powerpoint": "pptx",
        "ppt": "pptx",
    }.get(app_name, "xlsx")
    stamp = time.strftime("%Y%m%d_%H%M%S")
    millis = int((time.time() * 1000) % 1000)
    return os.path.abspath(f"{app_name}_{stamp}_{millis:03d}.{ext}")


def _action_names(actions):
    return {
        str(action.get("action", "")).strip().lower()
        for action in (actions or [])
        if isinstance(action, dict)
    }


def _is_fresh_file_intent(app_name, command_text, actions):
    action_names = _action_names(actions)
    create_actions = {
        "excel": {"create_workbook"},
        "word": {"create_document"},
        "powerpoint": {"create_presentation"},
        "ppt": {"create_presentation"},
    }
    if action_names & create_actions.get(app_name, set()):
        return True

    text = (command_text or "").lower()
    creation_words = ("create", "new", "start", "make")
    target_words = ("file", "workbook", "document", "presentation", "ppt")
    return any(w in text for w in creation_words) and any(w in text for w in target_words)


def _should_start_fresh(app_name, command_text, actions, file_path):
    if file_path:
        return False
    if _extract_named_file_path(command_text, app_name):
        return False

    open_actions = {
        "excel": {"open_workbook"},
        "word": {"open_document"},
        "powerpoint": {"open_presentation"},
        "ppt": {"open_presentation"},
    }
    return not bool(_action_names(actions) & open_actions.get(app_name, set()))


def _ensure_fresh_file_action(app_name, command_text, actions, file_path):
    actions = list(actions or [])
    if not actions or not _should_start_fresh(app_name, command_text, actions, file_path):
        return actions

    create_action = {
        "excel": "create_workbook",
        "word": "create_document",
        "powerpoint": "create_presentation",
        "ppt": "create_presentation",
    }.get(app_name)
    if not create_action:
        return actions
    if str(actions[0].get("action", "")).strip().lower() == create_action:
        return actions

    logging.info(f"Prepending {create_action} for fresh {app_name} file: {command_text}")
    return [{"action": create_action}, *actions]


def _resolve_output_file_path(app_name, command_text, actions, file_path):
    explicit = (file_path or "").strip()
    if explicit:
        return os.path.abspath(explicit)

    named = _extract_named_file_path(command_text, app_name)
    if named:
        # For "create/new file" style commands, avoid reusing locked/existing targets.
        if _is_fresh_file_intent(app_name, command_text, actions):
            return _next_available_path(named)
        return named

    if _should_start_fresh(app_name, command_text, actions, ""):
        return _generate_new_output_path(app_name)

    return ""


def _office_dependency_error(app_name):
    module_name, package_name = OFFICE_DEPENDENCIES.get(app_name, (None, None))
    if not module_name:
        return None
    try:
        __import__(module_name)
        return None
    except ModuleNotFoundError:
        return (
            f"{app_name.title()} support requires `{package_name}`. "
            f"Install it with `pip install {package_name}` or `pip install -r requirements.txt`."
        )


def _has_explicit_save_action(app_name, actions, command_text="", file_path=""):
    names = _action_names(actions)
    save_map = {
        "excel": {"save_workbook", "save_workbook_as"},
        "word": {"save_document", "save_document_as"},
        "powerpoint": {"save_presentation", "save_presentation_as"},
        "ppt": {"save_presentation", "save_presentation_as"},
    }
    if names & save_map.get(app_name, set()):
        return True

    # Treat an explicit target filename/path as save intent.
    if (file_path or "").strip():
        return True
    if _extract_named_file_path(command_text, app_name):
        return True
    return False


def _run_office_actions(app_name, actions, file_path=None, command_text=""):
    app_name = (app_name or "").lower().strip()
    output_path = (file_path or "").strip() or OFFICE_OUTPUTS.get(app_name, "output.xlsx")
    output_path = os.path.abspath(output_path)
    executed = []
    failures = []
    opened = False
    persisted = False
    dependency_error = _office_dependency_error(app_name)
    should_save = _has_explicit_save_action(
        app_name,
        actions,
        command_text=command_text,
        file_path=file_path or "",
    )

    if dependency_error:
        failures.append(dependency_error)
        return {
            "ok_count": 0,
            "total": len(actions),
            "executed": executed,
            "failures": failures,
            "output_path": output_path,
            "opened": opened,
            "dependency_error": dependency_error,
        }

    if app_name == "excel":
        from openpyxl import Workbook, load_workbook
        wb = load_workbook(output_path) if os.path.exists(output_path) else Workbook()
        ws = wb.active
        setattr(wb, "_path", output_path)
        executor = ExcelExecutor(wb, ws)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok: executed.append(action_name)
            else: failures.append(f"{action_name} failed")
        final_wb = getattr(executor, "wb", wb)
        setattr(final_wb, "_path", output_path)
        if should_save:
            try:
                final_wb.save(output_path)
                persisted = True
            except PermissionError:
                fallback_path = _next_available_path(output_path)
                final_wb.save(fallback_path)
                output_path = fallback_path
                persisted = True
                logging.warning(f"Excel target was locked. Saved to fallback path: {output_path}")
    elif app_name == "word":
        from docx import Document
        doc = Document(output_path) if os.path.exists(output_path) else Document()
        setattr(doc, "_path", output_path)
        executor = WordExecutor(doc)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok: executed.append(action_name)
            else: failures.append(f"{action_name} failed")
        final_doc = getattr(executor, "doc", doc)
        setattr(final_doc, "_path", output_path)
        if should_save:
            try:
                final_doc.save(output_path)
                persisted = True
            except PermissionError:
                fallback_path = _next_available_path(output_path)
                final_doc.save(fallback_path)
                output_path = fallback_path
                persisted = True
                logging.warning(f"Word target was locked. Saved to fallback path: {output_path}")
    elif app_name in ("powerpoint", "ppt"):
        from pptx import Presentation
        prs = Presentation(output_path) if os.path.exists(output_path) else Presentation()
        setattr(prs, "_path", output_path)
        executor = PowerPointExecutor(prs)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok: executed.append(action_name)
            else: failures.append(f"{action_name} failed")
        final_prs = getattr(executor, "prs", prs)
        setattr(final_prs, "_path", output_path)
        if should_save:
            try:
                final_prs.save(output_path)
                persisted = True
            except PermissionError:
                fallback_path = _next_available_path(output_path)
                final_prs.save(fallback_path)
                output_path = fallback_path
                persisted = True
                logging.warning(f"PowerPoint target was locked. Saved to fallback path: {output_path}")
    else:
        failures.append(f"Unsupported app: {app_name}")

    if not failures and persisted and os.path.exists(output_path):
        try:
            opened = bool(system_core.open_path(output_path))
        except Exception:
            opened = False

    return {
        "ok_count": len(executed),
        "total": len(actions),
        "executed": executed,
        "failures": failures,
        "output_path": output_path,
        "persisted": persisted,
        "opened": opened,
    }


def _handle_global_command(raw_text):
    """Handles system-wide agent: <app>: <command> triggers."""
    try:
        app_name, command = _extract_office_agent_command(raw_text)
        if app_name and command:
            if app_name == "ppt":
                app_name = "powerpoint"
            cache_key, actions, source = _resolve_actions(app_name, command)
            if not actions:
                logging.warning(f"No office action match for global command: {app_name}: {command}")
                return
            file_path = _resolve_output_file_path(app_name, command, actions, "")
            actions = _ensure_fresh_file_action(app_name, command, actions, file_path)
            summary = _run_office_actions(app_name, actions, file_path=file_path, command_text=command)
            if summary["failures"] and cache_key:
                command_map.remove_action(app_name, cache_key)
            logging.info(
                f"Global office [{source}] {app_name}: {command} -> "
                f"{summary['ok_count']}/{summary['total']} | {summary['output_path']}"
            )
            if summary.get("persisted"):
                _safe_speak(f"Executed {summary['ok_count']} actions in {app_name}")
            else:
                _safe_speak(f"Executed {summary['ok_count']} actions in {app_name}, not saved")
            return

        txt = (raw_text or "").strip()
        low = txt.lower()
        if low.startswith("agent "):
            sys_cmd = txt[len("agent "):].strip()
            sys_cmd = sys_cmd.replace("  ", " ").strip(" .,:;!?")
            if sys_cmd.startswith(("open ", "launch ", "start ", "run ", "boot ")):
                success, message = system_core.find_and_launch(sys_cmd)
                _safe_speak(
                    f"Opening {system_core.normalize_app_name(sys_cmd)}"
                    if success else f"Could not open {sys_cmd}"
                )
                logging.info(f"Voice system open [{sys_cmd}] => {success}: {message}")
            elif sys_cmd.startswith(("close ", "shut ", "exit ")):
                success, message = system_core.close_app(sys_cmd)
                _safe_speak(
                    f"Closing {system_core.normalize_app_name(sys_cmd)}"
                    if success else f"Could not close {sys_cmd}"
                )
                logging.info(f"Voice system close [{sys_cmd}] => {success}: {message}")
    except Exception as e:
        logging.error(f"Global command error: {e}\n{traceback.format_exc()}")


# Patch the keyboard/voice listener callbacks now that _handle_global_command is defined
_keyboard_listener.on_command = _handle_global_command
if _voice_listener:
    _voice_listener.on_command = _handle_global_command


# ===========================================================================
# ROUTES
# ===========================================================================

@app.route("/")
def index():
    return render_template("index.html")


# ---- System commands ------------------------------------------------------

@app.route("/execute", methods=["POST"])
def execute():
    try:
        data = request.json
        cmd = data.get("command", "").lower().strip()
        logging.info(f"Command: {cmd}")

        if cmd.startswith(("close ", "shut ", "exit ")):
            app_name = cmd.replace("close ", "").replace("shut ", "").replace("exit ", "").strip()
            success, message = system_core.close_app(app_name)
            _safe_speak(f"Closing {app_name}" if success else f"Could not close {app_name}")
            return jsonify(status="success" if success else "fail", message=message)

        app_name = system_core.normalize_app_name(cmd)
        success, message = system_core.find_and_launch(app_name)
        if success:
            _safe_speak(f"Opening {app_name}")
            return jsonify(status="success", message=message)

        _safe_speak(f"I couldn't find {app_name}. Please select it manually.")
        path = ui.manual_selector()
        if path:
            norm_app = system_core.normalize_app_name(app_name)
            config.save_memory(norm_app, path, is_store_app=False)
            launched = system_core.open_path(path)
            if launched:
                _safe_speak("Path saved. Opening now.")
                return jsonify(status="success", message="Manual Selection Saved")
            return jsonify(status="fail", message="Saved path, but launch failed")

        return jsonify(status="fail", message="Cancelled")

    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# ---- Office Agent ---------------------------------------------------------

@app.route("/office/execute", methods=["POST"])
def office_execute():
    try:
        return _office_execute_impl(request.json or {})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


def _office_execute_impl(data):
    app_name = (data.get("app") or "").lower().strip()
    command = (data.get("raw") or "").strip()
    full = (data.get("command") or "").strip()
    file_path = (data.get("file_path") or data.get("file") or "").strip()

    if not command and full:
        parsed_app, parsed_command = _extract_office_agent_command(full)
        if parsed_app and not app_name:
            app_name = (parsed_app or "").strip()
        if parsed_command:
            command = (parsed_command or "").strip()
        elif app_name:
            command = full

    if app_name == "ppt":
        app_name = "powerpoint"

    if app_name not in OFFICE_APPS or not command:
        return jsonify(status="fail", message="Missing/invalid app or command")

    cache_key, actions, source = _resolve_actions(app_name, command)
    if not actions:
        return jsonify(status="fail", message="No matching office command found. Try a more specific action like 'create a new workbook' or 'add heading Introduction'.", source=source)

    file_path = _resolve_output_file_path(app_name, command, actions, file_path)
    actions = _ensure_fresh_file_action(app_name, command, actions, file_path)

    summary = _run_office_actions(app_name, actions, file_path=file_path, command_text=command)
    if summary.get("dependency_error"):
        return jsonify(
            status="fail",
            message=summary["dependency_error"],
            source=source,
            output_file=summary["output_path"]
        )
    if summary["failures"] and cache_key:
        command_map.remove_action(app_name, cache_key)
        return jsonify(
            status="fail",
            message=f"✅ {summary['ok_count']}/{summary['total']} done | ❌ {' | '.join(summary['failures'])}",
            source=source,
            output_file=summary["output_path"]
        )

    return jsonify(
        status="success",
        message=(
            f"✅ Executed {summary['ok_count']} actions. Output: {summary['output_path']}"
            if summary.get("persisted")
            else f"✅ Executed {summary['ok_count']} actions (not saved). Add an explicit save command to write a file."
        ),
        source=source,
        output_file=summary["output_path"],
        persisted=summary.get("persisted", False),
        opened=summary.get("opened", False)
    )


@app.route("/command", methods=["POST"])
def office_command():
    try:
        return _office_execute_impl(request.json or {})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# ---- Voice control --------------------------------------------------------

@app.route("/voice/status", methods=["GET"])
def voice_status():
    if not _voice_listener:
        return jsonify(
            status="fail", available=False, enabled=False,
            message="Voice module unavailable. Install SpeechRecognition + PyAudio."
        )
    heard = _voice_listener.last_heard
    if time.time() - (_voice_listener.last_heard_at or 0) > 8:
        heard = ""
    return jsonify(
        status="success",
        available=_voice_listener.available,
        enabled=_voice_listener.is_running,
        armed=_voice_listener.armed,
        armed_seconds=round(_voice_listener.armed_seconds_left, 1),
        heard=heard,
        error=_voice_listener.last_error
    )


@app.route("/voice/start", methods=["POST"])
def voice_start():
    if not _voice_listener:
        return jsonify(status="fail", message="Voice module unavailable")
    ok = _voice_listener.start()
    voice_state["enabled"] = bool(ok)
    return jsonify(
        status="success" if ok else "fail",
        message="Voice listener started" if ok else (_voice_listener.last_error or "Could not start voice listener")
    )


@app.route("/voice/stop", methods=["POST"])
def voice_stop():
    if not _voice_listener:
        return jsonify(status="fail", message="Voice module unavailable")
    _voice_listener.stop()
    voice_state["enabled"] = False
    return jsonify(status="success", message="Voice listener stopped")


# ---- OCR ------------------------------------------------------------------

@app.route("/ocr/snip", methods=["POST"])
def ocr_snip():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        ocr_utils.snip_queue.put("snip")
        try:
            path = ocr_utils.result_queue.get(timeout=60)
        except Exception:
            return jsonify(status="fail", message="Snip timed out")
        if not path:
            return jsonify(status="fail", message="Snip cancelled")
        text = ocr_utils.image_to_text(path)
        last_ocr["text"] = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/screenshot", methods=["POST"])
def ocr_screenshot():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        path = ocr_utils.capture_fullscreen()
        text = ocr_utils.image_to_text(path)
        last_ocr["text"] = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/file", methods=["POST"])
def ocr_file():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        path = ui.file_selector(
            "Select an Image File",
            [("Images", "*.png *.jpg *.jpeg *.bmp *.tiff"), ("All Files", "*.*")]
        )
        if not path:
            return jsonify(status="fail", message="No file selected")
        text = ocr_utils.image_to_text(path)
        last_ocr["text"] = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text, message=f"OCR complete — {len(text)} chars")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/read", methods=["POST"])
def ocr_read():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        threading.Thread(target=ocr_utils.speak_text, args=(text,), daemon=True).start()
        return jsonify(status="success", message="Speaking...")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/stop_read", methods=["POST"])
def ocr_stop_read():
    try:
        ocr_utils.stop_speaking()
        return jsonify(status="success", message="Stopped")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/poll", methods=["GET"])
def ocr_poll():
    try:
        if last_ocr.get("pending"):
            last_ocr["pending"] = False
            return jsonify(
                status="ready",
                text=last_ocr["text"],
                message=f"Hotkey OCR complete — {len(last_ocr['text'])} chars"
            )
        return jsonify(status="waiting")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/save_txt", methods=["POST"])
def ocr_save_txt():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        path = ocr_utils.save_as_txt(text)
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/save_pdf", methods=["POST"])
def ocr_save_pdf():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install fpdf2: pip install fpdf2")
        path = pdf_utils.create_report(text, title="OCR Result")
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/clipboard", methods=["POST"])
def ocr_clipboard():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        ocr_utils.copy_to_clipboard(text)
        return jsonify(status="success", message="Copied to clipboard")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# ---- PDF Reader -----------------------------------------------------------

@app.route("/reader/open", methods=["POST"])
def reader_open():
    try:
        if not READER_AVAILABLE:
            return jsonify(status="fail", message="PDF reader module not found")
        path = ui.file_selector("Select PDF to Read", [("PDFs", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        threading.Thread(target=pdf_reader.start_reading, args=(path, 0), daemon=True).start()
        time.sleep(0.5)
        return jsonify(status="success", message="Reading started", **pdf_reader.get_status())
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/reader/pause", methods=["POST"])
def reader_pause():
    try:
        pdf_reader.pause_reading()
        return jsonify(status="success", message="Paused", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/resume", methods=["POST"])
def reader_resume():
    try:
        pdf_reader.resume_reading()
        return jsonify(status="success", message="Resumed", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/stop", methods=["POST"])
def reader_stop():
    try:
        pdf_reader.stop_reading()
        return jsonify(status="success", message="Stopped")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/next", methods=["POST"])
def reader_next():
    try:
        pdf_reader.next_page()
        return jsonify(status="success", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/prev", methods=["POST"])
def reader_prev():
    try:
        pdf_reader.prev_page()
        return jsonify(status="success", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/speed", methods=["POST"])
def reader_speed():
    try:
        data = request.json
        pdf_reader.set_speed(data.get("speed", 150))
        return jsonify(status="success", message=f"Speed: {data.get('speed')} WPM")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/status", methods=["GET"])
def reader_status():
    try:
        return jsonify(pdf_reader.get_status())
    except Exception:
        return jsonify(is_reading=False, is_paused=False, current_page=0, total_pages=0, speed=150)


# ---- PDF Tools ------------------------------------------------------------

@app.route("/pdf/merge", methods=["POST"])
def pdf_merge():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install pypdf: pip install pypdf")
        paths = pdf_utils.ask(
            kind="openmultiple",
            title="Select PDFs to Merge (hold Ctrl for multiple)",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if not paths:
            return jsonify(status="fail", message="No files selected.")
        out = pdf_utils.merge_pdfs(paths)
        if not out:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Merged {len(paths)} PDFs → {out}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/pdf/split", methods=["POST"])
def pdf_split():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install pypdf: pip install pypdf")
        path = ui.file_selector("Select PDF to Split", [("PDFs", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected.")
        pages = pdf_utils.split_pdf(path)
        if not pages:
            return jsonify(status="fail", message="Save cancelled or no pages.")
        return jsonify(status="success", message=f"Split into {len(pages)} files")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/pdf/create", methods=["POST"])
def pdf_create():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install fpdf2: pip install fpdf2")
        data = request.json
        text = data.get("text", "").strip()
        title = (data.get("title", "Report") or "Report").strip()
        if not text:
            return jsonify(status="fail", message="No text provided")
        path = pdf_utils.create_report(text, title=title)
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"PDF saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# ---- PDF Editor -----------------------------------------------------------

@app.route("/editor/open", methods=["POST"])
def editor_open():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        path = ui.file_selector("Select PDF to Edit", [("PDF Files", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        data = pdf_editor.extract_pdf_text(path)
        if data.get("status") != "success":
            return jsonify(status="fail", message=data.get("message", "Failed to open PDF"))
        return jsonify(status="success", file_path=path, pages=data["pages"], total_pages=data["total_pages"])
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/render-page", methods=["POST"])
def editor_render_page():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data = request.json
        pdf_path = data.get("file_path")
        page_num = data.get("page_num", 0)
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.render_page_as_image(pdf_path, page_num)
        if result.get("status") != "success":
            return jsonify(status="fail", message=result.get("message", "Render failed"))
        return jsonify(status="success", **{k: v for k, v in result.items() if k != "status"})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/save", methods=["POST"])
def editor_save():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data = request.json
        pdf_path = data.get("file_path")
        edits = data.get("edits", [])
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.save_edited_pdf(pdf_path, edits)
        if result.get("status") != "success":
            return jsonify(status="fail", message=result.get("message", "Save failed"))
        return jsonify(status="success", message=result.get("message", "Saved successfully"))
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/detect-form", methods=["POST"])
def editor_detect_form():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        path = ui.file_selector("Select PDF", [("PDF Files", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        fields = pdf_editor.detect_form_fields(path)
        return jsonify(
            status="success",
            is_form=len(fields) > 0,
            field_count=len(fields),
            fields=list(fields.keys()),
            file_path=path
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/editor/fill-form", methods=["POST"])
def editor_fill_form():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data = request.json
        pdf_path = data.get("file_path")
        form_data = data.get("form_data", {})
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.fill_form(pdf_path, form_data)
        if result:
            return jsonify(status="success", message="Form saved successfully")
        return jsonify(status="fail", message="Save cancelled")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/editor/get-field-options", methods=["POST"])
def editor_get_field_options():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data = request.json
        pdf_path = data.get("file_path")
        field_name = data.get("field_name")
        options = pdf_editor.get_form_field_options(pdf_path, field_name)
        return jsonify(status="success", field_name=field_name, options=options)
    except Exception as e:
        return jsonify(status="fail", message=f"Error: {str(e)}")


# ===========================================================================
# ENTRY POINT
# ===========================================================================

if __name__ == "__main__":

    # ---- OCR snip overlay (must be on main thread) ------------------------
    if OCR_AVAILABLE:
        threading.Thread(target=ocr_utils.run_snip_overlay_main_thread, daemon=True).start()

    # ---- OCR keyboard hotkeys ---------------------------------------------
    if KEYBOARD_AVAILABLE and OCR_AVAILABLE:
        keyboard.add_hotkey(
            "ctrl+shift+s",
            lambda: threading.Thread(
                target=ocr_utils.trigger_snip_and_ocr, args=(last_ocr,), daemon=True
            ).start()
        )
        keyboard.add_hotkey(
            "ctrl+shift+f",
            lambda: threading.Thread(
                target=ocr_utils.trigger_screenshot_and_ocr, args=(last_ocr,), daemon=True
            ).start()
        )
        print("🔑  Ctrl+Shift+S → Snip OCR  |  Ctrl+Shift+F → Fullscreen OCR")

    # ---- Global Office Agent listeners ------------------------------------
    threading.Thread(target=_clipboard_listener.start, daemon=True, name="ClipboardListener").start()
    threading.Thread(target=_keyboard_listener.start, daemon=True, name="KeyboardListener").start()
    print("⌨️   Global agent listener active")
    print("     Type  agent: excel: <command>  anywhere + Enter")

    if _voice_listener and _voice_listener.available:
        if _voice_listener.start():
            voice_state["enabled"] = True
            print("Voice wake listener active (say: agent <app> <command>)")
        else:
            print(f"Voice listener not started: {_voice_listener.last_error}")

    # ---- Start Flask ------------------------------------------------------
    flask_thread = threading.Thread(
        target=lambda: app.run(host="127.0.0.1", port=5000, debug=False),
        daemon=True
    )
    flask_thread.start()
    time.sleep(1)

    # ---- Open browser -----------------------------------------------------
    webbrowser.open("http://127.0.0.1:5000")
    print("✅  Agent running at http://127.0.0.1:5000")

    # ---- Dialog listener must be on main thread ---------------------------
    if PDF_AVAILABLE:
        pdf_utils.run_dialog_listener()
    else:
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n👋 Agent stopped.")
